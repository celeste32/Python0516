# -*- coding: utf-8 -*-
"""
Created on Sun May 16 14:08:29 2021

@author: user
"""

import openpyxl
from openpyxl.styles import Font
wb = openpyxl.Workbook() #建立一個Excel試算表
wb.create_sheet() #建立一個工作表
print(wb.sheetnames)
wb.create_sheet(index=0,title='學生')
wb.save('school.xlsx')

ws=wb['學生']
ws['A1']='姓名'
ws['B1']='性別'
ws['C1']='年齡'
ws['A2']='林小姐'
ws['B2']='女'
ws['C2']=30
peo = ['陳先生','男',26]
ws.append(peo)
data = [['王先生','男','24'],['謝小姐','女','34']]
for row in data:
    ws.append(row)
font = Font(name='微軟正黑體',size=25,bold=True)
ws['A1'].font = font
wb.save('school.xlsx')
