# -*- coding: utf-8 -*-
"""
Created on Sun May 16 13:25:15 2021

@author: user
"""

import openpyxl
fn = 'sales.xlsx'   #data_only=True只要值，不要公式
wb = openpyxl.load_workbook(fn,data_only=True)
ws = wb.active #預設工作表
ws = wb['2020Q3']
print(ws.title)
print(ws['A1'].value)
print(ws['B2'].value)
print(ws['A2'].value)
print(ws['E3'].value)
for i in range(1,ws.max_column+1):
    print(ws.cell(column=i,row=8).value,end='\t')
print()

for w in range(3,8):
    for s in range(1,6):
        print('%6s'%ws.cell(column=s,row=w).value,end='\t')
print()

for row in ws.columns: #把"欄"的數據轉化為"列"顯示
    for r in row:
        print(r.value)